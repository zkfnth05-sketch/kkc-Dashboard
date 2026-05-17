
import pandas as pd
import openpyxl

file_path = '4.4 도그쇼 A링 출진리스트.xlsx'

def analyze_excel(path):
    print(f"--- Analyzing {path} ---")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    
    # 1. 컬럼 너비 확인
    for col in range(1, 6):
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"Column {col_letter} width: {ws.column_dimensions[col_letter].width}")
    
    # 2. 병합된 셀 정보
    print("\nMerged Cells:")
    for merged_range in ws.merged_cells.ranges:
        print(f" - {merged_range}")
    
    # 3. 데이터 및 스타일 샘플 (상위 20행)
    print("\nRow Data & Basic Styles:")
    for row in range(1, 21):
        row_data = []
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            bold = "B" if cell.font and cell.font.bold else ""
            color = cell.fill.start_color.index if cell.fill and cell.fill.start_color else ""
            row_data.append(f"[{val}]({bold}|{color})")
        print(f"Row {row:02}: {' | '.join(row_data)}")

analyze_excel(file_path)
